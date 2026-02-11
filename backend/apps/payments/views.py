"""
ViewSets para Pagos
"""
from rest_framework import viewsets, permissions, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum
from django.utils import timezone
from datetime import timedelta
from .models import Payment, Invoice
from .serializers import PaymentSerializer, PaymentCreateSerializer, InvoiceSerializer


class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['payment_date', 'amount']
    ordering = ['-payment_date']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return PaymentCreateSerializer
        return PaymentSerializer
    
    def get_queryset(self):
        user = self.request.user
        queryset = Payment.objects.select_related('member__user', 'membership')
        
        if hasattr(user, 'member_profile') and not user.is_staff:
            queryset = queryset.filter(member__user=user)
        
        # Filtrar por estado
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Estadísticas de pagos"""
        today = timezone.now().date()
        month_start = today.replace(day=1)
        
        # Pagos del mes
        month_payments = Payment.objects.filter(
            payment_date__date__gte=month_start,
            status='completed'
        )
        
        total_month = month_payments.aggregate(total=Sum('amount'))['total'] or 0
        count_month = month_payments.count()
        
        # Pagos de hoy
        today_payments = Payment.objects.filter(
            payment_date__date=today,
            status='completed'
        )
        total_today = today_payments.aggregate(total=Sum('amount'))['total'] or 0
        
        return Response({
            'month': {
                'total': total_month,
                'count': count_month
            },
            'today': {
                'total': total_today,
                'count': today_payments.count()
            }
        })
    
    @action(detail=False, methods=['get'])
    def chart_data(self, request):
        """Datos para gráficas del dashboard"""
        from django.db.models import Sum, Count
        from datetime import timedelta
        from dateutil.relativedelta import relativedelta
        
        today = timezone.now().date()
        six_months_ago = today - relativedelta(months=5)
        
        monthly_revenue = []
        monthly_labels = []
        
        for i in range(6):
            month_date = six_months_ago + relativedelta(months=i)
            month_start = month_date.replace(day=1)
            month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
            if month_end > today:
                month_end = today
            
            revenue = Payment.objects.filter(
                payment_date__gte=month_start,
                payment_date__lte=month_end,
                status='completed'
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            monthly_revenue.append(float(revenue))
            monthly_labels.append(month_start.strftime('%b %Y'))
        
        three_months_ago = today - relativedelta(months=3)
        payment_methods = Payment.objects.filter(
            payment_date__gte=three_months_ago,
            status='completed'
        ).values('payment_method').annotate(
            count=Count('id')
        ).order_by('-count')
        
        method_labels = []
        method_values = []
        
        for pm in payment_methods:
            method_labels.append(dict(Payment.PAYMENT_METHODS).get(pm['payment_method'], pm['payment_method']))
            method_values.append(pm['count'])
        
        return Response({
            'revenue': {
                'labels': monthly_labels,
                'revenues': monthly_revenue
            },
            'payment_methods': {
                'labels': method_labels,
                'values': method_values
            }
        })
    
    @action(detail=False, methods=['get'])
    def my_payments(self, request):
        """Pagos del usuario actual (si es miembro)"""
        if not hasattr(request.user, 'member_profile'):
            return Response({'detail': 'No eres un miembro'}, status=400)
        
        payments = Payment.objects.filter(
            member=request.user.member_profile
        ).select_related('member__user', 'membership')
        
        serializer = self.get_serializer(payments, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Aprobar un pago pendiente"""
        payment = self.get_object()
        
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos'}, status=403)
        
        if payment.approve(request.user):
            return Response({
                'detail': 'Pago aprobado exitosamente',
                'payment': PaymentSerializer(payment).data
            })
        else:
            return Response({'detail': 'El pago no está pendiente'}, status=400)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Rechazar un pago pendiente"""
        payment = self.get_object()
        
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos'}, status=403)
        
        reason = request.data.get('reason', '')
        if not reason:
            return Response({'detail': 'Debes proporcionar un motivo'}, status=400)
        
        if payment.reject(reason, request.user):
            return Response({
                'detail': 'Pago rechazado',
                'payment': PaymentSerializer(payment).data
            })
        else:
            return Response({'detail': 'El pago no está pendiente'}, status=400)
    
    @action(detail=False, methods=['get'])
    def pending_count(self, request):
        """Contador de pagos pendientes"""
        if not request.user.is_staff:
            return Response({'count': 0})
        
        count = Payment.objects.filter(status='pending').count()
        return Response({'count': count})
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def export_report(self, request):
        """Exportar reporte de pagos a Excel"""
        # Verificar que el usuario sea staff
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos'}, status=403)
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        
        # Obtener filtros de fecha
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Filtrar pagos
        queryset = Payment.objects.select_related('member__user', 'membership__plan').order_by('-payment_date')
        
        if start_date:
            queryset = queryset.filter(payment_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(payment_date__lte=end_date)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Pagos"
        
        # Estilos
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = ['ID', 'Fecha', 'Miembro', 'Plan', 'Monto', 'Método', 'Estado', 'Referencia']
        ws.append(headers)
        
        # Aplicar estilo a headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        total = 0
        for payment in queryset:
            ws.append([
                payment.id,
                payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
                payment.member.user.get_full_name(),
                payment.membership.plan.name if payment.membership else 'N/A',
                float(payment.amount),
                payment.get_payment_method_display(),
                payment.get_status_display(),
                payment.reference_number or ''
            ])
            if payment.status == 'completed':
                total += float(payment.amount)
        
        # Totales
        ws.append([])
        total_row = ws.max_row + 1
        ws.append(['', '', '', 'TOTAL', total, '', '', ''])
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'reporte_pagos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response


class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer
    permission_classes = [permissions.IsAuthenticated]
