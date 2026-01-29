"""
ViewSets for Audit API
"""
from rest_framework import viewsets, permissions, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from .models import AuditLog, UserSession
from .serializers import (
    AuditLogSerializer, AuditLogListSerializer,
    UserSessionSerializer, AuditStatsSerializer
)


class IsAdminOrStaff(permissions.BasePermission):
    """Only allow access to admin and staff users."""
    
    def has_permission(self, request, view):
        return request.user and (request.user.is_staff or request.user.is_superuser)


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing audit logs.
    Read-only - logs cannot be modified.
    """
    
    queryset = AuditLog.objects.select_related('user').all()
    permission_classes = [IsAdminOrStaff]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['timestamp', 'action', 'model_name']
    ordering = ['-timestamp']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return AuditLogListSerializer
        return AuditLogSerializer
    
    def get_queryset(self):
        queryset = AuditLog.objects.select_related('user').all()
        
        # Filter by user
        user_id = self.request.query_params.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by action
        action = self.request.query_params.get('action')
        if action:
            queryset = queryset.filter(action=action)
        
        # Filter by model
        model_name = self.request.query_params.get('model')
        if model_name:
            queryset = queryset.filter(model_name=model_name)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(timestamp__gte=start_date)
        
        end_date = self.request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(timestamp__lte=end_date)
        
        # Filter by success/failure
        success = self.request.query_params.get('success')
        if success is not None:
            queryset = queryset.filter(success=success.lower() == 'true')
        
        # Filter by object
        object_id = self.request.query_params.get('object_id')
        if object_id and model_name:
            queryset = queryset.filter(object_id=object_id, model_name=model_name)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get audit statistics for dashboard."""
        now = timezone.now()
        today = now.date()
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)
        
        # Total counts
        total_today = AuditLog.objects.filter(timestamp__date=today).count()
        total_week = AuditLog.objects.filter(timestamp__gte=week_ago).count()
        total_month = AuditLog.objects.filter(timestamp__gte=month_ago).count()
        
        # Actions by type (last 30 days)
        actions_by_type = dict(
            AuditLog.objects.filter(timestamp__gte=month_ago)
            .values('action')
            .annotate(count=Count('id'))
            .values_list('action', 'count')
        )
        
        # Top users (last 30 days)
        top_users = list(
            AuditLog.objects.filter(timestamp__gte=month_ago, user__isnull=False)
            .values('user__first_name', 'user__last_name', 'user__email')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        
        # Recent critical actions (APPROVE, REJECT, DELETE)
        recent_critical = AuditLogListSerializer(
            AuditLog.objects.filter(
                action__in=['APPROVE', 'REJECT', 'DELETE', 'FREEZE']
            ).order_by('-timestamp')[:20],
            many=True
        ).data
        
        # Failed logins today
        failed_logins_today = AuditLog.objects.filter(
            timestamp__date=today,
            action='FAILED_LOGIN'
        ).count()
        
        # Active sessions
        active_sessions = UserSession.objects.filter(is_active=True).count()
        
        stats_data = {
            'total_today': total_today,
            'total_week': total_week,
            'total_month': total_month,
            'actions_by_type': actions_by_type,
            'top_users': top_users,
            'recent_critical': recent_critical,
            'failed_logins_today': failed_logins_today,
            'active_sessions': active_sessions
        }
        
        serializer = AuditStatsSerializer(stats_data)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export audit logs to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        # Styles
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = ['ID', 'Usuario', 'Acción', 'Modelo', 'Objeto', 'Timestamp', 'IP', 'Éxito']
        ws.append(headers)
        
        # Apply style to headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for log in queryset:
            ws.append([
                log.id,
                log.user.get_full_name() if log.user else 'Sistema',
                log.get_action_display(),
                log.model_name,
                log.object_repr,
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.ip_address or '',
                'Sí' if log.success else 'No'
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response


class UserSessionViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing user sessions."""
    
    queryset = UserSession.objects.select_related('user').all()
    serializer_class = UserSessionSerializer
    permission_classes = [IsAdminOrStaff]
    ordering = ['-login_time']
    
    def get_queryset(self):
        queryset = UserSession.objects.select_related('user').all()
        
        # Filter by user
        user_id = self.request.query_params.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by active status
        is_active = self.request.query_params.get('active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active sessions."""
        active_sessions = UserSession.objects.filter(is_active=True).select_related('user')
        serializer = self.get_serializer(active_sessions, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def terminate(self, request, pk=None):
        """Force terminate a session (admin only)."""
        if not request.user.is_superuser:
            return Response(
                {'detail': 'Solo administradores pueden terminar sesiones'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        session = self.get_object()
        session.logout_time = timezone.now()
        session.is_active = False
        session.save()
        
        return Response({'detail': 'Sesión terminada exitosamente'})
